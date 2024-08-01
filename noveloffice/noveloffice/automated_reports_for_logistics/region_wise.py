
import frappe
import pandas as pd
from frappe.utils.file_manager import save_file, get_file
import openpyxl
from openpyxl.styles import PatternFill, Font
import os

def create_region_report():
	current_date = frappe.utils.getdate()
	month = frappe.utils.format_date(current_date, 'MMM')
	year = current_date.year
	existing_doc = frappe.get_list('Region Wise Sales Custom Report', filters={'month': str(month), 'date': str(current_date)})
	
	if existing_doc:
		region_doc = frappe.get_doc('Region Wise Sales Custom Report', f'{str(month)}-{str(current_date)}')
	else:
		region_doc = frappe.new_doc('Region Wise Sales Custom Report')
		region_doc.month = month
		region_doc.date = current_date
	
	from_date = frappe.utils.get_first_day(current_date)
	to_date = frappe.utils.get_last_day(current_date)
	from_so_date = frappe.utils.getdate('2023-04-01')
	data = frappe.db.sql(f'''
		SELECT 
			combined.customer_name,
			SUM(combined.total_sales_sum) AS sales_amount,
			SUM(combined.total_debit_sum) AS monthly_collection,
			SUM(combined.amount_receivable) AS amount_receivable,
			SUM(combined.ar_overdue) AS ar_overdue,
			SUM(combined.so_abstract) AS so_abstract,
			SUM(combined.purchase) AS purchase,
			rwsct.region,
			CASE
			WHEN SUM(combined.total_sales_sum) != 0 THEN 
				((SUM(combined.total_sales_sum) - SUM(combined.purchase)) / SUM(combined.total_sales_sum))
			ELSE 0
			END AS margin
		FROM (
			-- Sales data with debit as 0
			SELECT 
				customer_name, 
				SUM(total_sale) AS total_sales_sum,
				0 AS total_debit_sum,
				0 AS amount_receivable,
				0 AS ar_overdue,
				0 AS so_abstract,
				0 AS purchase
			FROM `tabAs per Sales Child Table`
			WHERE parent = "{from_date}-{to_date}"
			GROUP BY customer_name

			UNION ALL

			-- Debit data with sales as 0
			SELECT 
				against_account AS customer_name, 
				0 AS total_sales_sum,
				SUM(debit) AS total_debit_sum,
				0 AS amount_receivable,
				0 AS ar_overdue,
				0 AS so_abstract,
				0 AS purchase
			FROM `tabCollections Report Child Table`
			WHERE parent = "{from_date}-{to_date}"
			GROUP BY against_account

			UNION ALL

			-- AR Summary data
			SELECT
				customer AS customer_name,
				0 AS total_sales_sum,
				0 AS total_debit_sum, 
				overdue_nondue AS amount_receivable,
				overdue AS ar_overdue,
				0 AS so_abstract,
				0 AS purchase
			FROM `tabCustom AR Summary child table`
			WHERE parent = "{current_date}"
			
			UNION ALL
			
			-- Sales Order Abstract data
			SELECT 
				customer AS customer_name,
				0 AS total_sales_sum,
				0 AS total_debit_sum, 
				0 AS amount_receivable,
				0 AS ar_overdue,
				SUM(open_so_amount) AS so_abstract,
				0 AS purchase
			FROM `tabSales Order Abstract Child Table`
			WHERE parent = "{from_so_date}-{current_date}"
			GROUP BY customer_name

			UNION ALL
			-- Sales Order Abstract data
			SELECT 
				customer_name AS customer_name,
				0 AS total_sales_sum,
				0 AS total_debit_sum, 
				0 AS amount_receivable,
				0 AS ar_overdue,
				0 AS so_abstract,
				SUM(purchase_amount) AS purchase
			FROM `tabGross Margin Child Table`
			WHERE parent = "{from_so_date}-{current_date}"
			AND MONTH(invoice_date) = MONTH(CURDATE()) -- Ensure purchase is in the current month
			AND YEAR(invoice_date) = YEAR(CURDATE())
			GROUP BY customer_name
		) AS combined
		LEFT JOIN `tabRegion Wise Supplier Child Table` AS rwsct
		ON combined.customer_name = rwsct.customer
		GROUP BY combined.customer_name, rwsct.region
		ORDER BY 
			CASE WHEN combined.customer_name = 'Total' THEN 0 ELSE 1 END,
			rwsct.region, 
			combined.customer_name;
	''', as_dict=True)
	
	region_doc.set('report', {})
	region_doc.set('report', data)
	region_doc.save()

	current_doc = frappe.get_doc('Region Wise Sales Custom Report', f'{month}-{current_date}')
	report_val = []
	for item in current_doc.report:
		report_val.append({'customer_name': item.customer_name,
						   'Purchase': item.purchase,
						   'Margin': float(item.margin),
						   'Sales Amount': item.sales_amount,
						   'Monthly Collection with GST': item.monthly_collection,
						   'Amount Receivable': item.amount_receivable,
						   'AR Overdue': item.ar_overdue,
						   'SO Abstract': item.so_abstract,
						   'Region': item.region})
	
	df = pd.DataFrame(report_val)
	df = df.sort_values(by=['Region'], key=lambda x: x=='Others', kind='stable')
	
	regions = df['Region'].dropna().unique()  # Remove None or NaN from unique regions
	
	# Create a list to store individual region DataFrames
	region_dfs = []

	for region in regions:
		region_df = df[df['Region'] == region].drop(columns=['Region'])
		region_df.columns = [f'{region}' if col == 'customer_name' else f'{col}' for col in region_df.columns]  # Prefix with region name
		region_dfs.append(region_df.reset_index(drop=True))
	# Determine the maximum number of rows needed
	max_rows = max([df.shape[0] for df in region_dfs])

	# Pad DataFrames with NaNs to ensure they have the same number of rows
	for i in range(len(region_dfs)):
		if region_dfs[i].shape[0] < max_rows:
			pad_rows = max_rows - region_dfs[i].shape[0]
			region_dfs[i] = pd.concat([region_dfs[i], pd.DataFrame([[pd.NA]*region_dfs[i].shape[1]]*pad_rows, columns=region_dfs[i].columns)], ignore_index=True)

	# Concatenate the DataFrames side by side
	final_df = pd.concat(region_dfs, axis=1)
	# final_df.insert(0, 'Region', [''] * len(final_df))
	# totals = final_df.select_dtypes(include=[float, int]).sum().to_dict()
	# final_df.loc[-1] = [final_df['Sales Amount'].sum()]
	# final_df['Sales Amount'] = final_df['Sales Amount'].sum()
	print(final_df)
	pd.options.display.float_format = '{:,.2f}'.format
	# print(final_df)
	# first_column = ['Region'] + [''] + [f'Month of {month}'] * (max_rows - 2)
	# final_df.insert(0, ' ', first_column)
	output_path = f'/home/frappe/frappe-bench/apps/erpnext/erpnext/selling/report/automated_reports_for_logistics/excel_sheets/Region Wise Report-{current_date}.xlsx'
	with pd.ExcelWriter(output_path) as writer:
		# Add the title row
		title_df = pd.DataFrame(['Region Wise Sales of Billing']).T
		title_df.to_excel(writer, sheet_name='Region Report', index=False, header=False, startrow=0)
		
		# Write the final DataFrame
		final_df.to_excel(writer, sheet_name='Region Report', index=False, startrow=1)

	# Apply styles using openpyxl
	wb = openpyxl.load_workbook(output_path)
	ws = wb['Region Report']

	# Fill the first row and columns with yellow color
	# yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
	# light_brown_fill = PatternFill(start_color='D3A77B', end_color='D3A77B', fill_type='solid')
	bold_font = Font(bold=True)

	ws.cell(row=1, column=1).font = bold_font

	for cell in ws[1]:
		# cell.fill = yellow_fill
		cell.font = bold_font

	
	for row in ws.iter_rows(min_row=1, max_row=max_rows + 1, min_col=1, max_col=1):
		for cell in row:
			# cell.fill = yellow_fill
			if cell.row == 1:
				cell.font = bold_font
	
	
	for col in ws.iter_cols(min_row=2, min_col=2, max_row=ws.max_row, max_col=ws.max_column):
		for cell in col:
			if isinstance(cell.value, (int, float)):
				if cell.column_letter in ['C', 'K', 'S', 'AA', 'AI']:  # Assuming 'H' is the column for margin
					cell.number_format = '0.00%'
				else:
					cell.number_format = '#,##0.00'
	# Save the workbook with styles
	wb.save(output_path)

	new_doc = frappe.new_doc('Logistics Report Attachments')
	new_doc.date = current_date
	new_doc.report_name = f'Region Wise Report-{current_date}'
	new_doc.insert()
	url = [{'file_url' : ''}]

	file_data = open(f'/home/frappe/frappe-bench/apps/erpnext/erpnext/selling/report/automated_reports_for_logistics/excel_sheets/Region Wise Report-{current_date}.xlsx', 'rb').read()
	file_attach = save_file(f'Region Wise Report-{current_date}.xlsx', file_data, 'Logistics Report Attachments', new_doc.name)
	file_list = frappe.get_list("File",
			filters = {
				'attached_to_name' : new_doc.name
			},
			as_list = True,
			)
	# print(file_list[0][0])
	file = frappe.get_doc("File",file_list[0][0])
	url[0]['file_url'] = file.file_url
	new_doc.attachment = file.file_url
	new_doc.save()
	print(url)
	recipient = ['orders@vibgyornet.com']
	# recipient = ['mohammed.s@noveloffice.in', 'ashish.k@vibgyornet.com']
	subject = f'Region Wise Report - {current_date}'
	message = '''Hello Team,<br>
	<br>
	PFA, for the updated Region Wise report as of 7:00 AM today.'''
	frappe.sendmail(
	recipients= recipient,
	bcc = ['mohammed.s@noveloffice.in'],
	subject=subject,
	message=message,
	attachments= url,
	expose_recipients= 'header',
	now = True
	)

	directory_path = f'/home/frappe/frappe-bench/apps/erpnext/erpnext/selling/report/automated_reports_for_logistics/excel_sheets/'

	if os.path.exists(directory_path):
		files = os.listdir(directory_path)
		for file_name in files:
			file_path = os.path.join(directory_path, file_name)
			os.remove(file_path)
